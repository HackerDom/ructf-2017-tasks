import openpyxl
import PIL.Image as Image
import PIL
import openpyxl.styles as styles
def create_document():
    work_book = openpyxl.Workbook()
    work_sheet = work_book.active
    image = Image.open("test.jpg")
    pixel_map = image.load()
    width, height = image.size
    for i in range(1,width):
        for j in range(1,height):
            cell = work_sheet.cell(row = j , column = i)
            pixel_color  =  pixel_map[i,j]
            color = "FF{0:02X}{1:02X}{2:02X}".format(*pixel_color)
            cell.fill = styles.PatternFill(fill_type="solid", start_color=color, end_color=color)
    work_book.save("test1.xlsx")

create_document()

